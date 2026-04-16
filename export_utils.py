import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from models import SelectedIndicator, ProjectInfo

_HEADERS = [
    "Organization",
    "Project",
    "Country",
    "Sector",
    "Indicator",
    "Definition",
    "Measurement Method",
    "Unit",
    "Frequency",
    "Source URL",
    "Data Source",
    "Framework Source",
    "Category",
    "Baseline",
    "Target",
]


def _row(si: SelectedIndicator, project_info: ProjectInfo) -> list[str]:
    return [
        project_info.organization_name,
        project_info.project_name,
        project_info.country,
        project_info.sector,
        si.indicator.title,
        si.indicator.definition,
        si.indicator.measurement_method,
        si.indicator.unit,
        si.effective_frequency,
        si.indicator.source_url,
        si.indicator.suggested_data_source,
        si.indicator.framework_source,
        si.indicator.category,
        si.baseline,
        si.target,
    ]


def export_to_csv(indicators: list[SelectedIndicator], project_info: ProjectInfo) -> str:
    """Return the M&E framework as a CSV string (UTF-8)."""
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow(_HEADERS)
    for si in indicators:
        writer.writerow(_row(si, project_info))
    return buf.getvalue()


def export_to_excel(indicators: list[SelectedIndicator], project_info: ProjectInfo) -> bytes:
    """Return the M&E framework as an Excel (.xlsx) byte string."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indicators"

    # Header styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    for si in indicators:
        ws.append(_row(si, project_info))

    # Auto-fit column widths (capped at 60)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Freeze the header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
